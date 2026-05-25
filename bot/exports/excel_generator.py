"""Excel market snapshot generator using openpyxl."""
from __future__ import annotations

import os
import tempfile
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from bot.utils.converters import (
    cents_per_lb_to_usd_per_kg,
    cents_per_lb_to_uzs_per_kg,
    usd_per_kg_to_uzs_per_kg,
    usd_per_kg_to_usd_per_lb,
)

_GREEN = "1A7A1A"
_WHITE = "FFFFFF"
_LIGHT = "F5F5F5"
_GREY_BORDER = Side(style="thin", color="CCCCCC")
_BORDER = Border(
    left=_GREY_BORDER, right=_GREY_BORDER, top=_GREY_BORDER, bottom=_GREY_BORDER
)
_HEADER_FILL = PatternFill("solid", fgColor=_GREEN)
_ALT_FILL = PatternFill("solid", fgColor=_LIGHT)
_HEADER_FONT = Font(name="Calibri", bold=True, color=_WHITE, size=10)
_BODY_FONT = Font(name="Calibri", size=10)
_TITLE_FONT = Font(name="Calibri", bold=True, color=_GREEN, size=14)
_SECTION_FONT = Font(name="Calibri", bold=True, color=_GREEN, size=11)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _write_header_row(ws, row: int, cols: list[str]) -> None:
    for c, text in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _write_data_row(ws, row: int, values: list, alt: bool = False) -> None:
    fill = _ALT_FILL if alt else None
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = _BODY_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
        if fill:
            cell.fill = fill


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(12, min(max_len + 4, 40))


def generate_excel(
    ice_data: Optional[dict],
    cotlook_data: Optional[dict],
    yarn_data: dict,
    fx_data: Optional[dict],
) -> str:
    """Generate Excel market snapshot. Returns path to temp file."""
    uzs_rate = fx_data["rate"] if fx_data else 12750.0
    ts = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Сводка"

    ws_sum.merge_cells("A1:F1")
    title_cell = ws_sum["A1"]
    title_cell.value = "CottonPulse — Снимок рынка"
    title_cell.font = _TITLE_FONT
    title_cell.alignment = _CENTER

    ws_sum.merge_cells("A2:F2")
    ts_cell = ws_sum["A2"]
    ts_cell.value = f"Сформирован: {ts}"
    ts_cell.font = Font(name="Calibri", italic=True, color="888888")
    ts_cell.alignment = _CENTER

    row = 4
    ws_sum.cell(row=row, column=1, value="Фьючерсы на хлопок").font = _SECTION_FONT
    row += 1

    _write_header_row(ws_sum, row, ["Инструмент", "ц/фунт", "USD/кг", "сум/кг", "Изм. (ц)", "Расчётно?"])
    row += 1

    if ice_data:
        p = ice_data["price"]
        _write_data_row(ws_sum, row, [
            "ICE CT=F",
            round(p, 2),
            round(cents_per_lb_to_usd_per_kg(p), 4),
            round(cents_per_lb_to_uzs_per_kg(p, uzs_rate), 0),
            round(ice_data.get("change", 0), 2),
            "Нет",
        ])
        row += 1

    if cotlook_data:
        p = cotlook_data["price"]
        _write_data_row(ws_sum, row, [
            "Cotlook A",
            round(p, 2),
            round(cents_per_lb_to_usd_per_kg(p), 4),
            round(cents_per_lb_to_uzs_per_kg(p, uzs_rate), 0),
            round(cotlook_data.get("change", 0), 2),
            "Да" if cotlook_data.get("estimated") else "Нет",
        ], alt=True)
        row += 1

    row += 1
    ws_sum.cell(row=row, column=1, value="Рынки пряжи").font = _SECTION_FONT
    row += 1

    _write_header_row(ws_sum, row, ["Страна", "Счёт", "USD/кг", "сум/кг", "USD/фунт", "Расчётно?"])
    row += 1

    country_labels = {"china": "Китай", "india": "Индия", "pakistan": "Пакистан"}
    for alt_idx, (country, data) in enumerate(yarn_data.items()):
        if data:
            p = data["price"]
            _write_data_row(ws_sum, row, [
                country_labels.get(country, country),
                data.get("count", "30s"),
                round(p, 2),
                round(usd_per_kg_to_uzs_per_kg(p, uzs_rate), 0),
                round(usd_per_kg_to_usd_per_lb(p), 4),
                "Да" if data.get("estimated") else "Нет",
            ], alt=alt_idx % 2 == 1)
            row += 1

    row += 1
    ws_sum.cell(row=row, column=1, value="Валютный курс").font = _SECTION_FONT
    row += 1

    _write_header_row(ws_sum, row, ["Пара", "Курс (сум)", "Изменение", "Изм. %", "Источник", ""])
    row += 1

    if fx_data:
        _write_data_row(ws_sum, row, [
            "USD/UZS",
            round(fx_data["rate"], 2),
            round(fx_data.get("change", 0), 2),
            f"{fx_data.get('change_pct', 0):+.3f}%",
            fx_data.get("source", "CBU"),
            "",
        ])
        row += 1

    _auto_width(ws_sum)

    # ── Sheet 2: History ──────────────────────────────────────────────────────
    ws_hist = wb.create_sheet("История цен")
    from bot.state import price_history, Commodity as C

    ws_hist.cell(row=1, column=1, value="История цен (последние 24ч)").font = _TITLE_FONT
    ws_hist.merge_cells("A1:D1")
    ws_hist["A1"].alignment = _CENTER

    commodity_map = {
        C.ICE_COTTON: "ICE Хлопок (ц/фунт)",
        C.COTLOOK: "Cotlook A (ц/фунт)",
        C.YARN_CHINA: "Пряжа Китай (USD/кг)",
        C.YARN_INDIA: "Пряжа Индия (USD/кг)",
        C.YARN_PAKISTAN: "Пряжа Пакистан (USD/кг)",
        C.USD_UZS: "USD/UZS",
    }

    col_offset = 1
    for commodity, label in commodity_map.items():
        points = list(price_history.get(commodity, []))
        if not points:
            continue
        ws_hist.cell(row=3, column=col_offset, value=label).font = _SECTION_FONT
        ws_hist.cell(row=4, column=col_offset, value="Время").font = _HEADER_FONT
        ws_hist.cell(row=4, column=col_offset).fill = _HEADER_FILL
        ws_hist.cell(row=4, column=col_offset).border = _BORDER
        ws_hist.cell(row=4, column=col_offset + 1, value="Цена").font = _HEADER_FONT
        ws_hist.cell(row=4, column=col_offset + 1).fill = _HEADER_FILL
        ws_hist.cell(row=4, column=col_offset + 1).border = _BORDER

        for r_idx, point in enumerate(points[-48:], 5):  # last 48 readings
            ws_hist.cell(row=r_idx, column=col_offset, value=point.timestamp.strftime("%H:%M")).font = _BODY_FONT
            ws_hist.cell(row=r_idx, column=col_offset).border = _BORDER
            ws_hist.cell(row=r_idx, column=col_offset + 1, value=round(point.price, 2)).font = _BODY_FONT
            ws_hist.cell(row=r_idx, column=col_offset + 1).border = _BORDER

        col_offset += 3

    _auto_width(ws_hist)

    # Save to temp file
    fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="cottonpulse_")
    os.close(fd)
    wb.save(path)
    return path
